"""Parse Vocabulary.xlsx (GRE curated list) into structured rows."""

from __future__ import annotations

XLSX_SOURCE = "Vocabulary.xlsx"

import re
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

import openpyxl


def default_xlsx_path() -> Path:
    base = Path(__file__).resolve().parents[3]
    for candidate in (
        base / "data" / "Vocabulary.xlsx",
        base.parent / "Vocabulary.xlsx",
    ):
        if candidate.is_file():
            return candidate
    return base.parent / "Vocabulary.xlsx"


@dataclass
class XlsxWordRow:
    lemma: str
    pos: str
    phonetic: Optional[str]
    vi_gloss: Optional[str]
    definition_en: str
    example: Optional[str]
    season: str
    order: int


def _map_pos(raw: Optional[str]) -> str:
    if not raw:
        return "noun"
    s = str(raw).strip().lower()
    if s.startswith("n") or "danh từ" in s:
        return "noun"
    if s.startswith("v") or "động từ" in s:
        return "verb"
    if "adj" in s or "tính từ" in s:
        return "adj"
    if "adv" in s or "trạng từ" in s:
        return "adv"
    return "noun"


def _first_example(raw: Optional[str]) -> Optional[str]:
    if not raw:
        return None
    text = str(raw).strip()
    if not text:
        return None
    parts = re.split(r"\n+", text)
    for part in parts:
        part = re.sub(r"^\d+\s*", "", part).strip()
        if len(part) > 12:
            return part
    return text[:500] if text else None


def _lemma_from_cell(word: Optional[str]) -> Optional[str]:
    if not word:
        return None
    lemma = str(word).strip().lower()
    lemma = lemma.split("\n")[0].strip()
    if not lemma or lemma in ("words", "word"):
        return None
    return lemma


def load_vocabulary_xlsx(path: Optional[Path] = None) -> List[XlsxWordRow]:
    xlsx_path = path or default_xlsx_path()
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"Vocabulary.xlsx not found at {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    rows: List[XlsxWordRow] = []
    seen: set[tuple[str, str]] = set()

    season_order = [
        "Season 1",
        "Season 2",
        "Season 3",
        "Season 4",
        "Season 5",
        "Season 6",
        "Season 7",
        "New words",
    ]

    for season in season_order:
        if season not in wb.sheetnames:
            continue
        ws = wb[season]
        is_new = season == "New words"
        order = 0
        for r in range(2, ws.max_row + 1):
            word_cell = ws.cell(r, 2).value
            lemma = _lemma_from_cell(word_cell)
            if not lemma:
                continue
            pos = _map_pos(ws.cell(r, 3).value if not is_new else None)
            key = (lemma, pos)
            if key in seen:
                continue
            seen.add(key)

            if is_new:
                meaning = str(ws.cell(r, 3).value or "").strip()
                definition_en = meaning.split("\n")[0].strip() or lemma
                rows.append(
                    XlsxWordRow(
                        lemma=lemma,
                        pos=pos,
                        phonetic=None,
                        vi_gloss=None,
                        definition_en=definition_en,
                        example=None,
                        season=season,
                        order=order,
                    )
                )
            else:
                vi = ws.cell(r, 5).value
                defn = ws.cell(r, 6).value
                rows.append(
                    XlsxWordRow(
                        lemma=lemma,
                        pos=pos,
                        phonetic=str(ws.cell(r, 4).value or "").strip() or None,
                        vi_gloss=str(vi).strip() if vi else None,
                        definition_en=str(defn).strip() if defn else lemma,
                        example=_first_example(ws.cell(r, 7).value),
                        season=season,
                        order=order,
                    )
                )
            order += 1
    wb.close()
    return rows


def gre_tier_for_season(season: str) -> str:
    if season in ("Season 1", "Season 2", "New words"):
        return "hard"
    if season in ("Season 3", "Season 4"):
        return "medium"
    return "medium"
